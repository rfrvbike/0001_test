from __future__ import annotations

import csv
import os
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterator, Protocol


REQUIRED_COLUMNS = [
    "post_text",
    "status",
    "scheduled_date",
    "posted_at",
    "tweet_id",
    "error",
]
POSTABLE_STATUSES = {"", "pending", "retry"}
NON_CANDIDATE_STATUSES = {
    "posted",
    "skipped",
    "content_error",
    "system_error",
    "error",
}


class QueueError(ValueError):
    """Raised when the post queue cannot be safely processed."""


class ContentRowError(ValueError):
    """Raised when only one row needs correction."""


@dataclass(frozen=True)
class PostRow:
    index: int
    csv_line_number: int
    post_text: str
    status: str
    scheduled_date: str
    posted_at: str
    tweet_id: str
    error: str


class PostQueue(Protocol):
    path: Path

    def read(self) -> tuple[list[dict[str, str]], list[str]]:
        ...

    def assert_writable(self) -> None:
        ...

    def write(self, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
        ...


class CsvPostQueue:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def read(self) -> tuple[list[dict[str, str]], list[str]]:
        if not self.path.exists():
            raise FileNotFoundError(f"Post queue not found: {self.path}")

        with self.path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            fieldnames = reader.fieldnames or []
            _validate_columns(fieldnames)
            rows = [{key: value or "" for key, value in row.items()} for row in reader]
        return rows, fieldnames

    def assert_writable(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(f"Post queue not found: {self.path}")
        temp_path = self.path.with_suffix(self.path.suffix + ".tmp")
        if temp_path.exists():
            raise QueueError(
                "Post queue temp file already exists. Close Excel/OneDrive sync, "
                f"inspect recovery state, then remove or recover: {temp_path}"
            )
        try:
            with self.path.open("r+b"):
                pass
            fd = os.open(str(temp_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("writable-check\n")
        except OSError as exc:
            raise QueueError(
                "Post queue is not writable. Close Excel and pause OneDrive sync "
                f"before posting: {self.path}"
            ) from exc
        finally:
            try:
                temp_path.unlink()
            except FileNotFoundError:
                pass

    def write(self, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
        _validate_columns(fieldnames)
        temp_path = self.path.with_suffix(self.path.suffix + ".tmp")
        try:
            with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
            os.replace(temp_path, self.path)
        except OSError as exc:
            raise QueueError(
                "Post queue update failed after preparing the temp file. Close Excel "
                f"and pause OneDrive sync, then recover from: {temp_path}"
            ) from exc


class OpenPyxlPostQueue:
    """Optional XLSX queue adapter.

    This adapter is used only when openpyxl already exists in the local
    environment. The project does not install it automatically.
    """

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def read(self) -> tuple[list[dict[str, str]], list[str]]:
        openpyxl = _load_openpyxl()
        if not self.path.exists():
            raise FileNotFoundError(f"Post queue not found: {self.path}")

        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook.active
        fieldnames = [str(cell.value or "").strip() for cell in sheet[1]]
        _validate_columns(fieldnames)

        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {}
            for index, name in enumerate(fieldnames):
                value = values[index] if index < len(values) else ""
                row[name] = "" if value is None else str(value)
            rows.append(row)
        return rows, fieldnames

    def assert_writable(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(f"Post queue not found: {self.path}")
        try:
            with self.path.open("r+b"):
                pass
        except OSError as exc:
            raise QueueError(
                "Post queue is not writable. Close Excel and pause OneDrive sync "
                f"before posting: {self.path}"
            ) from exc

    def write(self, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
        openpyxl = _load_openpyxl()
        _validate_columns(fieldnames)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(fieldnames)
        for row in rows:
            sheet.append([row.get(name, "") for name in fieldnames])
        workbook.save(self.path)


@contextmanager
def queue_lock(queue_path: str | Path) -> Iterator[Path]:
    """Create a sidecar lock file so two live runs cannot post the same row."""

    path = Path(queue_path)
    lock_path = path.with_suffix(path.suffix + ".lock")
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:
        raise QueueError(f"Queue is locked by another run: {lock_path}") from exc

    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(datetime.now().isoformat(timespec="seconds"))
        yield lock_path
    finally:
        try:
            lock_path.unlink()
        except FileNotFoundError:
            pass


def open_queue(path: str | Path) -> PostQueue:
    queue_path = Path(path)
    suffix = queue_path.suffix.lower()
    if suffix == ".csv" or queue_path.name.lower().endswith(".csv.example"):
        return CsvPostQueue(queue_path)
    if suffix == ".xlsx":
        return OpenPyxlPostQueue(queue_path)
    raise QueueError("Post queue must be a .csv or .xlsx file")


def assert_queue_writable(path: str | Path) -> None:
    open_queue(path).assert_writable()


def find_next_post(
    rows: list[dict[str, str]],
    *,
    today: date | None = None,
) -> PostRow | None:
    for candidate in iter_post_candidates(rows, today=today):
        try:
            validate_candidate_for_post(candidate, today=today)
        except ContentRowError:
            continue
        return candidate
    return None


def iter_post_candidates(
    rows: list[dict[str, str]],
    *,
    today: date | None = None,
) -> Iterator[PostRow]:
    effective_today = today or date.today()
    for index, row in enumerate(rows):
        text = (row.get("post_text") or "").strip()
        status = normalize_status(row.get("status", ""))
        scheduled_date = (row.get("scheduled_date") or "").strip()

        if status in NON_CANDIDATE_STATUSES:
            continue
        if status not in POSTABLE_STATUSES:
            continue

        try:
            scheduled = parse_scheduled_date(scheduled_date, row_error=True)
        except ContentRowError:
            scheduled = None
        else:
            if scheduled is not None and scheduled > effective_today:
                continue

        yield PostRow(
            index=index,
            csv_line_number=index + 2,
            post_text=text,
            status=status,
            scheduled_date=scheduled_date,
            posted_at=row.get("posted_at", ""),
            tweet_id=row.get("tweet_id", ""),
            error=row.get("error", ""),
        )


def validate_candidate_for_post(
    candidate: PostRow,
    *,
    today: date | None = None,
) -> None:
    if candidate.scheduled_date.strip():
        scheduled = parse_scheduled_date(candidate.scheduled_date, row_error=True)
        if scheduled is not None and scheduled > (today or date.today()):
            raise ContentRowError("scheduled_date is in the future")
    if not candidate.post_text.strip():
        raise ContentRowError("post_text is empty")
    if len(candidate.post_text.strip()) > 280:
        raise ContentRowError(
            f"post_text exceeds 280 characters: {len(candidate.post_text.strip())}"
        )


def mark_posted(
    rows: list[dict[str, str]],
    row_index: int,
    *,
    tweet_id: str,
    posted_at: datetime | None = None,
) -> None:
    rows[row_index]["status"] = "posted"
    rows[row_index]["posted_at"] = (posted_at or datetime.now()).isoformat(
        timespec="seconds"
    )
    rows[row_index]["tweet_id"] = tweet_id
    rows[row_index]["error"] = ""


def mark_error(rows: list[dict[str, str]], row_index: int, *, error: str) -> None:
    rows[row_index]["status"] = "error"
    rows[row_index]["error"] = error


def mark_content_error(
    rows: list[dict[str, str]],
    row_index: int,
    *,
    error: str,
) -> None:
    rows[row_index]["status"] = "content_error"
    rows[row_index]["error"] = error


def normalize_status(value: str | None) -> str:
    return (value or "").strip().lower()


def parse_scheduled_date(value: str, *, row_error: bool = False) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError as exc:
        error = f"scheduled_date must be YYYY-MM-DD, got: {value}"
        if row_error:
            raise ContentRowError(error) from exc
        raise QueueError(error) from exc


def _validate_columns(fieldnames: list[str]) -> None:
    missing = [name for name in REQUIRED_COLUMNS if name not in fieldnames]
    if missing:
        raise QueueError(f"Post queue is missing required columns: {missing}")


def _load_openpyxl():
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        raise QueueError(
            "XLSX queues require openpyxl to already be installed. "
            "Use CSV for the no-install dry-run path."
        ) from exc
    return openpyxl
